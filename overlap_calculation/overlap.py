
import numpy as np 
import openpyxl
from matplotlib import pyplot as plt

import shapely.geometry as sg
from shapely.affinity import rotate

def wafer_group( x, y ):
	# this function is for reading the GDS-based CSV files with x1 y1 and x2 y2 coordinates of the pores and slits  
	# assign x y coordinates to a group number (from 1 to 16, inclusive) based on the location on the wafer 
	# this function assumes that the center of rotation (point x=0,y=0) is somewhere around the center of the wafer
	#		hence the checks in the if statements for positive and negative x and y coordinates
	if x < -5e6: # nm 
		if y > 5e6:
			group = 1
		elif y > 0 and y < 5e6:
			group = 5
		elif y < 0 and y > -5e6:
			group = 9
		elif y < -5e6:
			group = 13
		else:
			ValueError
	elif x > -5e6 and x < 0: # nm 
		if y > 5e6:
			group = 2
		elif y > 0 and y < 5e6:
			group = 6
		elif y < 0 and y > -5e6:
			group = 10
		elif y < -5e6:
			group = 14
		else:
			ValueError
	elif x > 0 and x < 5e6: # nm 
		if y > 5e6:
			group = 3
		elif y > 0 and y < 5e6:
			group = 7
		elif y < 0 and y > -5e6:
			group = 11
		elif y < -5e6:
			group = 15
		else:
			ValueError
	elif x > 5e6: # nm 
		if y > 5e6:
			group = 4
		elif y > 0 and y < 5e6:
			group = 8
		elif y < 0 and y > -5e6:
			group = 12
		elif y < -5e6:
			group = 16
		else:
			ValueError
	else:
		ValueError
	return( group )

if __name__ == "__main__":

	# coordinates of the center of rotation/center of the wafer (from slits_corner_coords.xlsx file)

	center_x = 8478351.0 # nm 
	center_y = 7565000.977 # nm 

	# open and use Excel file that has coordinates (in nanometers) of bottom left and top right pore corners

	wb_slits = openpyxl.load_workbook("slits_corner_coords.xlsx")
	wsheet_s = wb_slits.active
	
	# preallocate the list of slits

	slits = [None]*16
	for sind in range(2,18):
		# initial slit coordinates - slits are revolving (from slit_corner_coords.xlsx file)
		slit_x1 = float(wsheet_s["A"+str(sind)].value) # nm
		slit_x2 = float(wsheet_s["C"+str(sind)].value) # nm
		slit_y1 = float(wsheet_s["B"+str(sind)].value) # nm
		slit_y2 = float(wsheet_s["D"+str(sind)].value) # nm
		# account for the center of rotation
		slit_x1 -= center_x
		slit_x2 -= center_x
		slit_y1 -= center_y
		slit_y2 -= center_y
		# store the slit in the list of slits
		index = wafer_group( slit_x1, slit_y1 ) 
		slits[index-1] = sg.box(slit_x1, slit_y1, slit_x2, slit_y2 ) 
		# plot 
		xs, ys = slits[index-1].exterior.xy 
		plt.plot(xs,ys,'b')
		# add textboxes in the plot
		if slit_x1 < -5e6:
			plt.text( 0.95*slit_x1, slit_y1, 'Group '+str(index)) 
		elif slit_x1 > -5e6 and slit_x1 < 0:
			plt.text( 0.9*slit_x1, slit_y1, 'Group '+str(index)) 
		elif slit_x1 > 0 and slit_x1 < 5e6:
			plt.text( 1.1*slit_x1, slit_y1, 'Group '+str(index)) 
		else:
			plt.text( 1.05*slit_x1, slit_y1, 'Group '+str(index)) 

	# open and use Excel file that has coordinates (in nanometers) of bottom left and top right pore corners

	wb_pores = openpyxl.load_workbook("pores_corner_coords.xlsx")
	wsheet_p = wb_pores.active

	# pore coordinates (from pores_corner_coords.xlsx file)

	pores = [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]] 
	for pind in range(2,578): 
		# pore extents on the x axis
		pore_x1 = float(wsheet_p["A"+str(pind)].value) # nm
		pore_x2 = float(wsheet_p["C"+str(pind)].value) # nm
		# pore extents on the y axis
		pore_y1 = float(wsheet_p["B"+str(pind)].value) # nm
		pore_y2 = float(wsheet_p["D"+str(pind)].value) # nm
		# account for the center of rotation
		pore_x1 -= center_x
		pore_x2 -= center_x
		pore_y1 -= center_y
		pore_y2 -= center_y
		# store the pore in the list of pores
		index = wafer_group( pore_x1, pore_y1 ) 
		pores[index-1].append(sg.box(pore_x1, pore_y1, pore_x2, pore_y2))
		# plot 
		xp, yp  = pores[index-1][-1].exterior.xy 
		plt.plot(xp,yp,'k')

	# some preliminaries for the output Excel file 
	
	wb_output = openpyxl.Workbook(write_only=True)
	ws_write = wb_output.create_sheet(0)

	header1 = ['Angle']	
	header2 = ['(degrees)']
	for i in range(16):
		header1.append('Group')
		header1.append(str(i+1))
		header2.append('area (nm^2)')
		header2.append('opc') # opc = open pore count (i.e. number of open pores) 

	ws_write.append(['Note 0: Group refers to one of 16 groups of slits and pores, top row has groups 1,2,3,4, second row has groups 5-8, and so on '])
	ws_write.append(['Note 1: negative Angle means clockwise rotation from vertical']) 
	ws_write.append(['Note 2: positive Angle means counter clockwise rotation from vertical'])
	ws_write.append(['Note 3: area (nm^2) means the total area within a group of pores that is open due to the overlap of slit and pore(s) '])
	ws_write.append(['Note 4: "opc" stands for "open pore count", i.e. the number of open pores that contribute to the area (nm^2) mentioned in Note 3 (see above)'])

	ws_write.append(header1)
	ws_write.append(header2)

	# rotation  - positive angle counter clockwise
	#			- negative angle clockwise
	#			approximately 1e-2 degree increments; for 601 angle values about 1.5 minutes of CPU time are required on Grigoriy's aging Core i5 laptop

	angle_start = -2.5
	angle_stop = 2.5
	angles = np.linspace(angle_start,angle_stop,601) # angles from negative to positive: counter clockwise rotation

	# obtain and write output data to the Excel file

	for angle in angles: 
		
		overlap_at_angle = [] 
		overlap_at_angle.append(angle)

		for cs, slit in enumerate(slits): 
			
			# rotate slit about the center of the wafer
			slit_rot = rotate(slit, angle, origin=(0,0)) 

			# check for open pore area within the pore group that corresponds to the slit
			open_pore_area = 0.0
			open_pores = 0
			for pore in pores[cs]: 
				if slit_rot.intersects(pore):
					# the area calculation is easy - the answer is already in nm^2 (square nanometers)
					open_pore_area += slit_rot.intersection(pore).area 
					open_pores += 1

			# store data for writing to Excel
			overlap_at_angle.append(np.round(open_pore_area))
			overlap_at_angle.append(open_pores)

			# plot the rotated slits, vary color based on rotation angle
			# rotation is counter clockwise 
			# 		starts at clockwise deviation (negative angle) from zero angle, goes to counter clockwise (positive angle) deviation
			xsr,ysr = slit_rot.exterior.xy
			# xsr and ysr contain x and y coordinates of the corners of the rotated shape
			#	start at bottom right corner, go counter clockwise (i.e up, left, down, right)
			#		for a rectangle, there will be 5 x values and 5 y values:
			#			enough information to draw 4 lines and form a closed shape (rectangle) 
			if angle < 0.0:
				colour = 'g' # start rotation at the green slit
			elif angle > 0.0:
				colour ='r' # finish rotation at the red slit
			elif angle == 0.0:
				colour = 'b' # zero rotation (aligned slits and pores) - blue slit
			else:
				ValueError

			# plot bottom left and top right corners (as dots) of all rotated slits
			# this will give a visual representation of slit motion trajectory 
			# 	we will see which pores are never opened by the slits
			if angle != angle_start and angle != angle_stop and angle != 0.0:
				# bottom left corner
				plt.plot( xsr[3], ysr[3], marker='.', color=colour ) 
				# top right corner 
				plt.plot( xsr[1], ysr[1], marker='.', color=colour )  
			else:
				# plot rotated slit
				plt.plot(xsr,ysr,colour) 

		# write to Excel spreadsheet
		ws_write.append(overlap_at_angle)
	
	# save the output Excel spreadsheet
	wb_output.save(filename = 'output_data_overlaps.xlsx')

	# display the final figure
	plt.axis('equal') 
	plt.show()

